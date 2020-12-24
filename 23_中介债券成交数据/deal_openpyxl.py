#!/usr/bin/env python
# coding: utf-8

# In[2]:


import pandas as pd
import numpy as np
import re



file_path=input('拖入文件后按回车，需和"页面模板.xlsx"在同一文件夹内：')
file_name=re.split(r'\|/|\\|//',file_path)[-1].split('.')[0]
print(file_name)
with open (file_path) as f:
    text=f.read()
    f.close()
# In[3]:



with open (file_path) as f:
    text=f.read()
    f.close()


# In[4]:


fenlei={'短融':[],
        '中票':[],
        '企业债':[],
        '其他':[]}


for i,line in enumerate(text.split('\n')):
    line=re.split(' +|\t+',line.strip())
    if len(line)==1 and line!=['']:
        if re.search('|'.join(['短融','中票','企业债','其他']) ,line[0])!=None:
            now_list=re.search('|'.join(['短融','中票','企业债','其他']) ,line[0]).group()
            
    
    if len (line)>=5:
        year=line[0]
        try:
#             rating=line[['AAA|(A\-1)' in i for i in line ].index(True)] # 找评级
            rating=line[[re.search('AAA|(A\-1)',i)!=None  for i in line].index(True)] #找评级

            ytm=line[[re.search('\d+\.\d+|\d+',i)!=None  for i in line[3:]].index(True)+3].replace('%','') #找ytm
            name=line[[re.search('[^\x00-\xff]+',i)!=None  for i in line[1:]].index(True)+1]#找名称
            fenlei[now_list].append([year,name,ytm,rating])

        except:
            pass
            # if now_list=='短融':
            #     print('无法获取信息',line)
            #
        
    else:
        pass
        

len(fenlei['短融'])


# In[5]:


fenlei_df ={i:pd.DataFrame(fenlei[i],columns='剩余期限 简称 收益率 评级'.split(' ')) for  i in fenlei.keys()}


# In[6]:


def get_day(days:str):
    main_day=days.split('+')[0]
    if 'D' in main_day:
        real_day=float(re.findall('\d+\.\d+|\d+',main_day)[0])
    elif 'Y' in main_day:
        real_day=float(re.findall('\d+\.\d+|\d+',main_day)[0])*365
    else:
        real_day=main_day
    return real_day


# In[7]:


for key in fenlei_df:
    df=fenlei_df[key]
    df['day']=df['剩余期限'].apply(lambda x:get_day(x))
    df.sort_values('day',inplace=True)
    df.index=[i for i in range(len(df))]
    fenlei_df[key]=df.drop('day',axis=1) #去掉辅助列
    
fenlei_df['中票'].head()


# ## 输出到excel中

# In[118]:


from shutil import copyfile
copyfile('./页面模板.xlsx','./{}.xlsx'.format(file_name))

import openpyxl
wb=openpyxl.load_workbook('{}.xlsx'.format(file_name))
sht=wb['Sheet1']


# In[119]:



def get_excel_coor(now_line):
    page=int(now_line/120)
    line=now_line%60
    if now_line%120<60:
        col='A'
    else:
        col='E'
    row=page*62+line+3
    return str(col)+str(row)
now_line=120
print(get_excel_coor(now_line))


# In[120]:


from openpyxl.styles import Font, colors, Alignment
def small_title(sheet,coor,title):
    end_coor_dic={'A':'D','E':'H'}
    hebing=coor+':'+end_coor_dic[coor[0]]+coor[1:]
    sheet.merge_cells(hebing)
    sheet[coor]=title
    sheet[coor].alignment = Alignment(horizontal='center', vertical='center')
    print(hebing,title)
small_title(sht,'A3','短融')


# In[121]:


# 取消合并所有单元格

for page in range(6):
    page_range='A{}:H{}'.format(page*62+3,page*62+62)
    try:
        sht.unmerge_cells(page_range)
    except:
        print('unmerged')
        pass
    # Borders(11) 内部垂直边线。    # Borders(12) 内部水平边线。

#     for line in [11,12]:
#         sht.c(page_range).api.Borders(line).LineStyle = 1
#         sht.range(page_range).api.Borders(line).Weight = 2
#         sht.range(page_range).api.HorizontalAlignment = -4131
#         sht.range(page_range).value=None
#         sht.range(page_range).api.NumberFormat = "@"
sht['E1']=file_name


# In[128]:


def write_area(sht,coor,write_df):
    if coor[0]=='A':
        min_col=1
        max_col=4
    if coor[0]=='E':
        min_col=5
        max_col=8
    min_row=int(coor[1:])
    max_row=min_row+len(write_df)-1
#     print(write_df)
    for i,row in enumerate(sht.iter_rows(min_row=min_row,min_col=min_col, max_col=max_col, max_row=max_row)):
        for j,cell in enumerate(row):
#             print(cell)
            try:
                cell.value=write_df.iloc[i,j]
            except:
                print(cell)


# In[129]:


now_line=0


def xiugai_df(write_df):
    buf=write_df.iloc[0,0]
    write_df.set_index('剩余期限',inplace=True)
    write_df.columns=write_df.iloc[0]
    write_df.index.name=buf
    write_df=write_df.iloc[1:] # 通过一个很愚蠢的方式去掉表头
    return write_df

for one in fenlei_df:
    # 输入标签
    coor=get_excel_coor(now_line)
    small_title(sht,coor,one)
    now_line+=1
    
    
    if len(fenlei_df[one])==0:
        continue
    
    line_left=60-now_line%60# 剩下多少行的空间
    small_df=fenlei_df[one]
    
    while len(small_df)>0:
#         print('还剩多少行:',line_left,'需要多少行:',len(small_df))
        coor=get_excel_coor(now_line) # 计算新的坐标位置
        print(coor)
        # 填充剩余位置
        write_df=small_df.iloc[:line_left]
        
        write_area(sht,coor,write_df)
        
        # 更新位置
        now_line+=len(write_df)
        
        small_df=small_df.iloc[line_left:]
        line_left=60-now_line%60  # 更新剩下多少行的空间
        


# In[131]:


wb.save('{}.xlsx'.format(file_name))


# In[ ]:




